import streamlit as st
import pandas as pd
import random
import time
import base64
from pathlib import Path

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Belanja & Menang Winner Selection",
    layout="centered",
)

# ── Image helpers ──────────────────────────────────────────────────────────────
def get_base64(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

def apply_background(image_path: str):
    ext = Path(image_path).suffix.lower().lstrip(".")
    mime = "jpeg" if ext in ("jpg", "jpeg") else ext
    b64 = get_base64(image_path)
    st.markdown(f"""
    <style>
    .stApp {{
        background-image: url("data:image/{mime};base64,{b64}");
        background-size: cover;
        background-position: center top;
        background-attachment: scroll;
        background-repeat: no-repeat;
    }}
    .stApp::before {{
        content: "";
        position: fixed;
        inset: 0;
        background: rgba(10, 10, 20, 0.40);
        z-index: 0;
    }}
    [data-testid="stAppViewContainer"] > * {{ position: relative; z-index: 1; }}
    </style>
    """, unsafe_allow_html=True)

# ── File names ─────────────────────────────────────────────────────────────────
BG_FILE      = "background.jpg"
LOGO_FILE    = "logo.png"
BACKUP_COUNT = 10   # number of backups drawn per prize in one press

if Path(BG_FILE).exists():
    apply_background(BG_FILE)
else:
    st.markdown("<style>.stApp { background: #0d0d14; }</style>", unsafe_allow_html=True)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700&family=DM+Sans:wght@400;500&display=swap');

  #MainMenu { visibility: hidden; }
  header[data-testid="stHeader"] { background: transparent !important; height: 0 !important; }
  footer { visibility: hidden; }
  [data-testid="stToolbar"] { display: none !important; }

  html, body, [class*="css"],
  p, span, div, label, li, a,
  h1, h2, h3, h4, h5, h6,
  .stMarkdown, .stText,
  [data-testid="stMarkdownContainer"] p,
  [data-testid="stMarkdownContainer"] span,
  [data-testid="stMarkdownContainer"] li,
  [data-testid="stWidgetLabel"] p,
  [data-testid="stWidgetLabel"] span,
  [data-testid="stFileUploaderLabel"] p,
  [data-testid="stFileUploaderLabel"] span,
  .stFileUploader label, .stFileUploader p,
  .stAlert p, [data-testid="stAlert"] p, [data-testid="stAlert"] span {
    font-family: 'DM Sans', sans-serif !important;
    color: #ffffff !important;
  }
  h1, h2, h3 { font-family: 'Playfair Display', serif !important; color: #ffffff !important; }

  [data-testid="stAlert"] { background: rgba(22,22,42,0.85) !important; border-radius: 10px !important; }

  [data-testid="stAppViewContainer"] { padding-top: 0 !important; }
  [data-testid="block-container"] {
    padding-top: 0 !important;
    padding-left: 2rem !important;
    padding-right: 2rem !important;
    max-width: 680px !important;
  }

  .logo-topright { position: fixed; top: 16px; right: 24px; z-index: 999; }
  .logo-topright img { width: 120px; object-fit: contain; }

  .title-block {
    position: fixed; top: 150px; right: 40px; z-index: 998;
    text-align: right; max-width: 320px;
  }
  .title-block h1 {
    font-family: 'Playfair Display', serif !important;
    font-size: 32px !important; font-weight: 700 !important;
    color: #ffffff !important; line-height: 1.3 !important;
    margin: 0 0 8px !important; text-shadow: 2px 2px 10px rgba(0,0,0,0.95);
  }
  .title-block p {
    font-size: 14px !important; color: #eeeeee !important;
    margin: 0 !important; text-shadow: 1px 1px 6px rgba(0,0,0,0.95);
  }

  [data-testid="stFileUploadDropzone"] {
    border: 1.5px solid #cccccc !important;
    border-radius: 12px !important;
    background: rgba(255,255,255,0.88) !important;
  }

  [data-testid="stFileUploadDropzone"] p,
  [data-testid="stFileUploadDropzone"] span,
  [data-testid="stFileUploaderDropzoneInstructions"] span,
  [data-testid="stFileUploaderDropzoneInstructions"] p,
  [data-testid="stFileUploaderDropzoneInstructions"] div {
    color: #222222 !important;
  }

  [data-testid="stFileUploadDropzone"] svg {
    fill: #333333 !important;
  }

  /* Streamlit 1.55 upload button fix: hide duplicated built-in text and draw one clean label */
  [data-testid="stFileUploadDropzone"] button {
    position: relative !important;
    width: 150px !important;
    height: 56px !important;
    overflow: hidden !important;
    white-space: nowrap !important;
    color: transparent !important;
    border: 2px solid #f5a3ad !important;
    border-radius: 14px !important;
    background: rgba(255,255,255,0.25) !important;
  }

  [data-testid="stFileUploadDropzone"] button * {
    display: none !important;
    visibility: hidden !important;
    color: transparent !important;
  }

  [data-testid="stFileUploadDropzone"] button::after {
    content: "Upload" !important;
    position: absolute !important;
    inset: 0 !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 16px !important;
    font-weight: 500 !important;
    color: #ffffff !important;
  }

  /* Winner card */
  .tier-card {
    border-radius: 14px; padding: 1.2rem 1.5rem; margin-bottom: 0.75rem;
    display: flex; align-items: center; gap: 16px;
  }
  .tier-gold   { background: rgba(59,45,14,0.92);  border: 1px solid #c9a227; }
  .tier-silver { background: rgba(30,30,40,0.92);  border: 1px solid #8e9db5; }
  .tier-bronze { background: rgba(42,22,16,0.92);  border: 1px solid #a0522d; }
  .tier-icon { font-size: 40px; line-height: 1; flex-shrink: 0; }
  .tier-text { flex: 1; }
  .tier-label {
    font-size: 11px; font-weight: 500; letter-spacing: .12em;
    text-transform: uppercase; opacity: .75; margin-bottom: 3px; color: #ffffff !important;
  }
  .tier-winner { font-family: 'Playfair Display', serif !important; font-size: 24px; font-weight: 700; }

  /* Backup card — shows list of 10 */
  .backup-card {
    background: rgba(10,10,25,0.88); border: 1px solid #333355;
    border-radius: 14px; padding: 1rem 1.5rem; margin-bottom: 0.75rem;
  }
  .backup-card-title {
    font-size: 12px; font-weight: 600; letter-spacing: .1em;
    text-transform: uppercase; color: #aaaacc !important;
    margin-bottom: 0.75rem; padding-bottom: 0.5rem;
    border-bottom: 1px solid #333355;
  }
  .backup-grid {
    display: grid; grid-template-columns: 1fr 1fr; gap: 4px 16px;
  }
  .backup-item {
    display: flex; align-items: center; gap: 8px;
    padding: 4px 0; font-size: 14px; color: #ffffff !important;
  }
  .backup-num {
    font-size: 11px; color: #666688 !important;
    min-width: 18px; text-align: right; flex-shrink: 0;
  }

  /* Next prize box */
  .next-prize-box {
    background: rgba(10,10,25,0.80); border: 1px solid #444466;
    border-radius: 12px; padding: 1.2rem 1.5rem; margin-bottom: 1rem;
    display: flex; align-items: center; gap: 14px;
  }
  .next-prize-box .np-icon { font-size: 48px; line-height: 1; flex-shrink: 0; }
  .next-prize-box .np-info { flex: 1; }
  .next-prize-box .np-label { font-size: 11px; text-transform: uppercase; letter-spacing: .1em; color: #aaaacc !important; margin-bottom: 4px; }
  .next-prize-box .np-title { font-family: 'Playfair Display', serif; font-size: 26px; font-weight: 700; color: #ffffff !important; }
  .next-prize-box .np-pool  { font-size: 12px; color: #aaaacc !important; margin-top: 4px; }

  .round-banner {
    background: rgba(74,63,165,0.85); border: 1px solid #6a5fcc;
    border-radius: 10px; padding: 0.6rem 1.2rem;
    font-size: 13px; font-weight: 500; color: #ffffff !important;
    display: inline-block; margin-bottom: 1rem;
  }

  /* Rolling animation */
  .rolling-text {
    font-family: 'Playfair Display', serif !important;
    font-size: 26px; font-weight: 700; color: #ffffff !important;
    animation: pulse 0.35s infinite alternate; text-align: center; padding: 1rem 0;
  }
  @keyframes pulse { from { opacity:.3; } to { opacity:1; } }

  /* Summary */
  .summary-card {
    background: rgba(10,10,25,0.88); border: 1px solid #333355;
    border-radius: 16px; padding: 1.5rem; margin-bottom: 1rem;
  }
  .summary-title {
    font-family: 'Playfair Display', serif; font-size: 20px; font-weight: 700;
    color: #ffffff; margin: 0 0 1rem;
    border-bottom: 1px solid #333355; padding-bottom: 0.75rem;
  }
  .summary-winner-row {
    display: flex; align-items: center; gap: 14px;
    padding: 0.5rem 0; border-bottom: 1px solid rgba(255,255,255,0.08); margin-bottom: 0.75rem;
  }
  .summary-icon { font-size: 26px; flex-shrink: 0; width: 32px; text-align: center; }
  .summary-prize { font-size: 11px; text-transform: uppercase; letter-spacing: .1em; color: #aaaacc !important; margin-bottom: 2px; }
  .summary-name  { font-family: 'Playfair Display', serif; font-size: 20px; font-weight: 700; }
  .gold-name   { color: #f5c842 !important; }
  .silver-name { color: #c0cfe0 !important; }
  .bronze-name { color: #d4825a !important; }
  .summary-backup-label { font-size: 11px; text-transform: uppercase; letter-spacing: .1em; color: #aaaacc !important; margin-bottom: 0.5rem; }
  .summary-backup-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 4px 16px; }
  .summary-backup-item {
    font-size: 14px; color: #ccccdd !important;
    padding: 3px 0; display: flex; gap: 8px; align-items: center;
  }
  .summary-backup-num { color: #666688 !important; min-width: 18px; font-size: 11px; }

  hr { border-color: #444466 !important; }

  div[data-testid="stButton"] > button {
    background: rgba(10,10,20,0.6) !important; border: 1px solid #666688 !important;
    border-radius: 8px !important; color: #ffffff !important;
    font-family: 'DM Sans', sans-serif !important; font-weight: 500 !important; transition: all .18s !important;
  }
  div[data-testid="stButton"] > button:hover {
    border-color: #aaaacc !important; background: rgba(28,28,48,0.9) !important;
  }
  div[data-testid="stButton"] > button[kind="primary"] {
    background: #4a3fa5 !important; border-color: #6a5fcc !important;
  }
  div[data-testid="stButton"] > button[kind="primary"]:hover { background: #5a50be !important; }
</style>
""", unsafe_allow_html=True)

# ── Logo & Title ───────────────────────────────────────────────────────────────
if Path(LOGO_FILE).exists():
    logo_b64 = get_base64(LOGO_FILE)
    ext  = Path(LOGO_FILE).suffix.lower().lstrip(".")
    mime = "jpeg" if ext in ("jpg", "jpeg") else ext
    st.markdown(f'<div class="logo-topright"><img src="data:image/{mime};base64,{logo_b64}" alt="Logo"></div>', unsafe_allow_html=True)

st.markdown("""
<div class="title-block">
  <h1>Belanja & Menang<br>Winner Selection</h1>
  <p>Three prizes. One name each.<br>Let fate decide.</p>
</div>
""", unsafe_allow_html=True)

st.markdown("<div style='height: 240px;'></div>", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
def _init():
    defaults = {
        "names":        [],
        "pool":         [],
        "winners":      {"3rd": None, "2nd": None, "1st": None},
        "backups":      {"3rd": [],   "2nd": [],   "1st": []},
        "stage":        "upload",   # upload|w_ready|w_drawing|w_done|b_ready|b_drawing|summary
        "current_tier": "3rd",
        "draw_order":   ["3rd", "2nd", "1st"],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init()
s = st.session_state

TIER_META = {
    "3rd": {"label": "3rd Prize", "badge": "🥉", "cls": "tier-bronze", "name_color": "#d4825a"},
    "2nd": {"label": "2nd Prize", "badge": "🥈", "cls": "tier-silver", "name_color": "#c0cfe0"},
    "1st": {"label": "1st Prize", "badge": "🥇", "cls": "tier-gold",   "name_color": "#f5c842"},
}
NAME_CLS = {"3rd": "bronze-name", "2nd": "silver-name", "1st": "gold-name"}

def reset():
    for k in ["names","pool","winners","backups","stage","current_tier","_uploaded_raw","_uploaded_flag"]:
        if k in st.session_state:
            del st.session_state[k]
    _init()

def next_tier_after(tier):
    idx = s["draw_order"].index(tier)
    return s["draw_order"][idx + 1] if idx + 1 < len(s["draw_order"]) else None

def render_winner_card(tier_key, name):
    m = TIER_META[tier_key]
    st.markdown(f"""
    <div class="tier-card {m['cls']}">
      <div class="tier-icon">{m['badge']}</div>
      <div class="tier-text">
        <div class="tier-label">{m['label']}</div>
        <div class="tier-winner" style="color:{m['name_color']} !important;">{name}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

def render_backup_card(tier_key):
    m    = TIER_META[tier_key]
    bkps = s["backups"][tier_key]
    if not bkps:
        return
    items_html = "".join([
        f'<div class="backup-item"><span class="backup-num">{i+1}.</span>{name}</div>'
        for i, name in enumerate(bkps)
    ])
    st.markdown(f"""
    <div class="backup-card">
      <div class="backup-card-title">🔄 Backup List — {m['label']} ({len(bkps)} names)</div>
      <div class="backup-grid">{items_html}</div>
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD
# ─────────────────────────────────────────────────────────────────────────────
if s["stage"] == "upload":
    uploaded = st.file_uploader("Drop the name list here", type=["csv"])
    if uploaded:
        try:
            import io
            raw = uploaded.read()
            df  = None
            for enc in ["utf-8", "utf-8-sig", "latin-1", "cp1252", "iso-8859-1"]:
                try:
                    df = pd.read_csv(io.BytesIO(raw), header=0, encoding=enc)
                    break
                except Exception:
                    continue
            if df is None:
                st.error("Could not read the file. Please save your CSV as UTF-8 and try again.")
                st.stop()

            col   = df.columns[0]
            names = df[col].dropna().astype(str).str.strip().tolist()
            names = [n for n in names if n]
            min_needed = 3 + (3 * BACKUP_COUNT)
            if len(names) < min_needed:
                st.error(f"Need at least {min_needed} participants (3 winners + {BACKUP_COUNT} backups × 3 prizes) — only found {len(names)}.")
            else:
                s["names"] = names
                s["pool"]  = names.copy()
                st.success(f"✓ Loaded **{len(names)} participants** successfully!")

                # ── Name list in scrollable box, left-aligned, fixed width ──
                numbered_items = "".join([
                    f'<div style="font-size:13px; padding:4px 0; border-bottom:1px solid rgba(255,255,255,0.07); color:#ffffff; display:flex; gap:10px;">' +
                    f'<span style="color:#666688; min-width:24px; text-align:right; flex-shrink:0;">{i+1}.</span>' +
                    f'<span>{n}</span></div>'
                    for i, n in enumerate(names)
                ])
                st.markdown(f"""
                <div style="
                    width: 320px;
                    background: rgba(10,10,25,0.85);
                    border: 1px solid #333355;
                    border-radius: 12px;
                    padding: 0.8rem 1rem;
                    height: 280px;
                    overflow-y: auto;
                    margin-bottom: 1rem;
                ">
                  <div style="font-size:11px; text-transform:uppercase; letter-spacing:.1em; color:#aaaacc; margin-bottom:0.5rem; font-weight:600;">
                    📋 Participant List ({len(names)} names)
                  </div>
                  {numbered_items}
                </div>
                """, unsafe_allow_html=True)

                if st.button("Start Lucky Draw →", type="primary", use_container_width=True):
                    s["stage"] = "w_ready"; s["current_tier"] = "3rd"
                    st.rerun()
        except Exception as e:
            st.error(f"Could not read file: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# WINNER ROUND — ready (wait for button click)
# ─────────────────────────────────────────────────────────────────────────────
elif s["stage"] == "w_ready":
    st.markdown('<div class="round-banner">🏆 Round 1 — Main Winners</div>', unsafe_allow_html=True)
    for t in s["draw_order"]:
        if s["winners"][t]:
            render_winner_card(t, s["winners"][t])

    tier = s["current_tier"]
    meta = TIER_META[tier]
    st.markdown(f"""
    <div class="next-prize-box">
      <div class="np-icon">{meta['badge']}</div>
      <div class="np-info">
        <div class="np-label">Up next</div>
        <div class="np-title">{meta['label']}</div>
        <div class="np-pool">{len(s['pool'])} names remaining in pool</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        if st.button(f"Draw {meta['label']} {meta['badge']}", type="primary", use_container_width=True):
            s["stage"] = "w_drawing"; st.rerun()
    with col2:
        if st.button("Reset", use_container_width=True):
            reset(); st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# WINNER ROUND — animate + pick winner
# ─────────────────────────────────────────────────────────────────────────────
elif s["stage"] == "w_drawing":
    st.markdown('<div class="round-banner">🏆 Round 1 — Main Winners</div>', unsafe_allow_html=True)
    for t in s["draw_order"]:
        if s["winners"][t]:
            render_winner_card(t, s["winners"][t])

    roll = st.empty()
    for _ in range(20):
        roll.markdown(f"<div class='rolling-text'>{random.choice(s['pool'])}</div>", unsafe_allow_html=True)
        time.sleep(0.07)

    winner = random.choice(s["pool"])
    # Remove ALL entries of this name from pool (duplicate participants)
    s["pool"] = [n for n in s["pool"] if n != winner]
    s["winners"][s["current_tier"]] = winner
    roll.empty()

    nxt = next_tier_after(s["current_tier"])
    s["current_tier"] = nxt if nxt else s["current_tier"]
    s["stage"]        = "w_ready" if nxt else "w_done"
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# WINNER ROUND — all 3 winners drawn
# ─────────────────────────────────────────────────────────────────────────────
elif s["stage"] == "w_done":
    st.markdown('<div class="round-banner">🏆 Round 1 — Main Winners</div>', unsafe_allow_html=True)
    for t in s["draw_order"]:
        render_winner_card(t, s["winners"][t])

    st.markdown("---")
    st.markdown(f"<p style='color:#cccccc; font-size:13px; text-align:center;'>All main winners drawn! Now draw {BACKUP_COUNT} backups per prize — 1 press per prize.</p>", unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        if st.button("Start Backup Draw →", type="primary", use_container_width=True):
            s["stage"] = "b_ready"; s["current_tier"] = "3rd"; st.rerun()
    with col2:
        if st.button("Reset", use_container_width=True):
            reset(); st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# BACKUP ROUND — ready (1 press = 10 backups)
# ─────────────────────────────────────────────────────────────────────────────
elif s["stage"] == "b_ready":
    tier = s["current_tier"]
    meta = TIER_META[tier]

    st.markdown('<div class="round-banner">🔄 Round 2 — Backup Winners</div>', unsafe_allow_html=True)

    # Show main winners
    for t in s["draw_order"]:
        render_winner_card(t, s["winners"][t])

    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

    # Show already completed backup lists
    for t in s["draw_order"]:
        if t == tier:
            break
        render_backup_card(t)

    # Next backup draw box
    st.markdown(f"""
    <div class="next-prize-box">
      <div class="np-icon">🔄</div>
      <div class="np-info">
        <div class="np-label">Backup — up next</div>
        <div class="np-title">{meta['label']} — {BACKUP_COUNT} names</div>
        <div class="np-pool">{len(s['pool'])} names remaining in pool</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        if st.button(f"Draw {BACKUP_COUNT} Backups — {meta['label']} {meta['badge']}", type="primary", use_container_width=True):
            s["stage"] = "b_drawing"; st.rerun()
    with col2:
        if st.button("Reset", use_container_width=True):
            reset(); st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# BACKUP ROUND — animate + pick 10 backups at once
# ─────────────────────────────────────────────────────────────────────────────
elif s["stage"] == "b_drawing":
    tier = s["current_tier"]
    meta = TIER_META[tier]

    st.markdown('<div class="round-banner">🔄 Round 2 — Backup Winners</div>', unsafe_allow_html=True)

    for t in s["draw_order"]:
        render_winner_card(t, s["winners"][t])

    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

    # Show already completed backup lists
    for t in s["draw_order"]:
        if t == tier:
            break
        render_backup_card(t)

    # Rolling animation — runs through names quickly
    roll = st.empty()
    for _ in range(25):
        roll.markdown(f"<div class='rolling-text'>{random.choice(s['pool'])}</div>", unsafe_allow_html=True)
        time.sleep(0.07)
    roll.empty()

    # Draw 10 unique backups at once (unique by name, not by entry)
    unique_pool = list(dict.fromkeys(s["pool"]))  # deduplicated, order preserved
    drawn_backups = random.sample(unique_pool, min(BACKUP_COUNT, len(unique_pool)))
    # Remove ALL entries of each drawn name from pool
    drawn_set = set(drawn_backups)
    s["pool"] = [n for n in s["pool"] if n not in drawn_set]
    s["backups"][tier] = drawn_backups

    # Move to next tier or summary
    nxt = next_tier_after(tier)
    s["current_tier"] = nxt if nxt else tier
    s["stage"]        = "b_ready" if nxt else "summary"
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY — winners + 10 backups each
# ─────────────────────────────────────────────────────────────────────────────
elif s["stage"] == "summary":
    st.markdown("<h3 style='text-align:center; color:#ffffff; margin-bottom:1.5rem;'>🎉 Final Results</h3>", unsafe_allow_html=True)

    for t in reversed(s["draw_order"]):   # 1st → 2nd → 3rd
        m      = TIER_META[t]
        winner = s["winners"][t]
        bkps   = s["backups"][t]
        nc     = NAME_CLS[t]

        backup_items_html = "".join([
            f'<div class="summary-backup-item"><span class="summary-backup-num">{i+1}.</span>{name}</div>'
            for i, name in enumerate(bkps)
        ])

        st.markdown(f"""
        <div class="summary-card">
          <div class="summary-title">{m['badge']} {m['label']}</div>
          <div class="summary-winner-row">
            <div class="summary-icon">🏆</div>
            <div>
              <div class="summary-prize">Winner</div>
              <div class="summary-name {nc}">{winner}</div>
            </div>
          </div>
          <div class="summary-backup-label">🔄 Backup List ({BACKUP_COUNT} names)</div>
          <div class="summary-backup-grid">{backup_items_html}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)

    # ── Export to Excel ──
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def build_excel():
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        tier_styles = {
            "1st": {"title": "1st Prize 🥇", "header_fill": "B8860B", "winner_fill": "FFF3CD", "backup_fill": "FFFDE7"},
            "2nd": {"title": "2nd Prize 🥈", "header_fill": "4A6B8A", "winner_fill": "E8F0F7", "backup_fill": "F0F6FF"},
            "3rd": {"title": "3rd Prize 🥉", "header_fill": "7B4A2D", "winner_fill": "F7EDE8", "backup_fill": "FFF5F0"},
        }

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for t in reversed(s["draw_order"]):  # 1st → 2nd → 3rd
            m      = TIER_META[t]
            style  = tier_styles[t]
            winner = s["winners"][t]
            bkps   = s["backups"][t]

            ws = wb.create_sheet(title=style["title"])
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 45

            # ── Title row ──
            ws.merge_cells("A1:B1")
            title_cell = ws["A1"]
            title_cell.value = style["title"]
            title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill      = PatternFill("solid", fgColor=style["header_fill"])
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # ── Winner section header ──
            ws.merge_cells("A2:B2")
            wh = ws["A2"]
            wh.value     = "🏆  WINNER"
            wh.font      = Font(bold=True, size=11, color="FFFFFF")
            wh.fill      = PatternFill("solid", fgColor="333355")
            wh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[2].height = 22

            # Winner name row
            ws["A3"].value = "1"
            ws["A3"].font      = Font(color="888888", size=10)
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A3"].fill      = PatternFill("solid", fgColor=style["winner_fill"])
            ws["A3"].border    = border

            ws["B3"].value     = winner
            ws["B3"].font      = Font(bold=True, size=12)
            ws["B3"].alignment = Alignment(vertical="center", indent=1)
            ws["B3"].fill      = PatternFill("solid", fgColor=style["winner_fill"])
            ws["B3"].border    = border
            ws.row_dimensions[3].height = 22

            # ── Blank spacer ──
            ws.row_dimensions[4].height = 8

            # ── Backup section header ──
            ws.merge_cells("A5:B5")
            bh = ws["A5"]
            bh.value     = f"🔄  BACKUP LIST  ({len(bkps)} names)"
            bh.font      = Font(bold=True, size=11, color="FFFFFF")
            bh.fill      = PatternFill("solid", fgColor="333355")
            bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[5].height = 22

            # Backup name rows
            for i, name in enumerate(bkps):
                row = 6 + i
                ws[f"A{row}"].value     = str(i + 1)
                ws[f"A{row}"].font      = Font(color="888888", size=10)
                ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"A{row}"].fill      = PatternFill("solid", fgColor=style["backup_fill"])
                ws[f"A{row}"].border    = border

                ws[f"B{row}"].value     = name
                ws[f"B{row}"].font      = Font(size=11)
                ws[f"B{row}"].alignment = Alignment(vertical="center", indent=1)
                ws[f"B{row}"].fill      = PatternFill("solid", fgColor=style["backup_fill"])
                ws[f"B{row}"].border    = border
                ws.row_dimensions[row].height = 20

        # ── Summary sheet ──
        ws_sum = wb.create_sheet(title="Summary", index=0)
        ws_sum.column_dimensions["A"].width = 12
        ws_sum.column_dimensions["B"].width = 40
        ws_sum.column_dimensions["C"].width = 10
        ws_sum.column_dimensions["D"].width = 40

        ws_sum.merge_cells("A1:D1")
        hdr = ws_sum["A1"]
        hdr.value     = "Belanja & Menang — Winner Selection Results"
        hdr.font      = Font(bold=True, size=14, color="FFFFFF")
        hdr.fill      = PatternFill("solid", fgColor="4a3fa5")
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 32

        headers = ["Prize", "Winner", "Backup #", "Backup Name"]
        fills   = ["D9D9D9"] * 4
        for col_idx, (h, f) in enumerate(zip(headers, fills), 1):
            c = ws_sum.cell(row=2, column=col_idx)
            c.value     = h
            c.font      = Font(bold=True, size=10)
            c.fill      = PatternFill("solid", fgColor="333355")
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center", indent=1)
            c.border    = border
        ws_sum.row_dimensions[2].height = 20

        current_row = 3
        for t in reversed(s["draw_order"]):
            m      = TIER_META[t]
            winner = s["winners"][t]
            bkps   = s["backups"][t]
            style  = tier_styles[t]
            start  = current_row

            # Winner row
            ws_sum.cell(row=current_row, column=1).value     = m["label"]
            ws_sum.cell(row=current_row, column=1).font      = Font(bold=True)
            ws_sum.cell(row=current_row, column=1).fill      = PatternFill("solid", fgColor=style["winner_fill"])
            ws_sum.cell(row=current_row, column=1).border    = border
            ws_sum.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

            ws_sum.cell(row=current_row, column=2).value     = winner
            ws_sum.cell(row=current_row, column=2).font      = Font(bold=True)
            ws_sum.cell(row=current_row, column=2).fill      = PatternFill("solid", fgColor=style["winner_fill"])
            ws_sum.cell(row=current_row, column=2).border    = border
            ws_sum.cell(row=current_row, column=2).alignment = Alignment(vertical="center", indent=1)

            ws_sum.cell(row=current_row, column=3).fill   = PatternFill("solid", fgColor=style["winner_fill"])
            ws_sum.cell(row=current_row, column=3).border = border
            ws_sum.cell(row=current_row, column=4).fill   = PatternFill("solid", fgColor=style["winner_fill"])
            ws_sum.cell(row=current_row, column=4).border = border
            ws_sum.row_dimensions[current_row].height = 20
            current_row += 1

            # Backup rows
            for i, name in enumerate(bkps):
                ws_sum.cell(row=current_row, column=1).fill   = PatternFill("solid", fgColor=style["backup_fill"])
                ws_sum.cell(row=current_row, column=1).border = border
                ws_sum.cell(row=current_row, column=2).fill   = PatternFill("solid", fgColor=style["backup_fill"])
                ws_sum.cell(row=current_row, column=2).border = border

                ws_sum.cell(row=current_row, column=3).value     = str(i + 1)
                ws_sum.cell(row=current_row, column=3).font      = Font(color="888888", size=10)
                ws_sum.cell(row=current_row, column=3).fill      = PatternFill("solid", fgColor=style["backup_fill"])
                ws_sum.cell(row=current_row, column=3).border    = border
                ws_sum.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")

                ws_sum.cell(row=current_row, column=4).value     = name
                ws_sum.cell(row=current_row, column=4).font      = Font(size=10)
                ws_sum.cell(row=current_row, column=4).fill      = PatternFill("solid", fgColor=style["backup_fill"])
                ws_sum.cell(row=current_row, column=4).border    = border
                ws_sum.cell(row=current_row, column=4).alignment = Alignment(vertical="center", indent=1)
                ws_sum.row_dimensions[current_row].height = 18
                current_row += 1

            # Merge prize label across its rows
            if current_row - start > 1:
                ws_sum.merge_cells(f"A{start}:A{current_row-1}")
                ws_sum[f"A{start}"].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1  # spacer row

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    excel_data = build_excel()
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📥 Export Results to Excel",
            data=excel_data,
            file_name="Belanja_Menang_Results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    with col2:
        if st.button("🔄 Start a new draw", use_container_width=True):
            reset(); st.rerun()
